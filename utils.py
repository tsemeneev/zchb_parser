import openpyxl
import pandas as pd
from openpyxl import load_workbook

from fake_useragent import UserAgent
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By


class Parser:
    def __init__(self):
        self.options = webdriver.ChromeOptions()
        # self.options.add_argument('--headless')
        self.options.add_experimental_option('excludeSwitches', ['enable-automation'])
        self.options.add_experimental_option('useAutomationExtension', False)
        self.options.add_argument(f'--user-agent={UserAgent.random}')
        self.options.add_argument('--disable-blink-features=AutomationControlled')
        self.options.add_argument('--disable-dev-shm-usage')
        self.options.add_argument('--disable-lazy-loading')
        self.inn_iter_count = 0
        self.proxy_list_counter = 0

    def del_humanity_check(self):
        driver = webdriver.Chrome(options=self.options)
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Promise;
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Object;
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Symbol;
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Function;
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Proxy;
                delete window.cdc_adoQpoasnfa76pfcZLmcfl_Array;
            """
        })
        return driver

    @staticmethod
    def check_box_detect(driver):
        check_box = driver.find_elements(By.ID, 'challenge-form')
        if len(check_box) > 0:
            return True
        return False

    def proxy_changer(self, count):
        proxy_list = [
            '127.0.0.1:8080',
            '127.0.0.1:8081',
        ]
        if count > 4:
            self.proxy_list_counter = 0
            count = 0
        self.options.add_argument(f'--proxy-server={proxy_list[count]}')
        driver = webdriver.Chrome(options=self.options)
        self.proxy_list_counter += 1
        return driver

    def parse(self):
        inn_list, company_name = self.get_inn_list()
        driver = self.del_humanity_check()
        driver.maximize_window()
        try:
            for count in range(len(inn_list) + int(len(inn_list) / 20)):
                driver.implicitly_wait(2)
                driver.get('https://zachestnyibiznes.ru/')

                if self.check_box_detect(driver):
                    driver = self.proxy_changer(self.proxy_list_counter)
                    continue

                search_box = driver.find_element(By.ID, 'autocomplete-0-input')
                search_box.clear()
                search_box.send_keys(inn_list[self.inn_iter_count])

                search_box.send_keys(Keys.ENTER)
                if self.check_box_detect(driver):
                    driver = self.proxy_changer(self.proxy_list_counter)
                    continue

                try:
                    button = driver.find_elements(By.CLASS_NAME, 'col-md-7')
                    if len(button) > 0:
                        button[0].find_element(By.TAG_NAME, 'a').click()
                        if self.check_box_detect(driver):
                            driver = self.proxy_changer(self.proxy_list_counter)
                            continue
                except Exception as e:
                    print(e)
                    continue

                driver.find_element(By.XPATH, '//span[@title="Посмотреть"]').click()
                if self.check_box_detect(driver):
                    driver = self.proxy_changer(self.proxy_list_counter)
                    continue

                phone = driver.find_elements(By.XPATH, '//a[contains(@href, "tel:")]')
                phone = 'Не указан' if len(phone) == 0 else phone[0].text
                email = driver.find_elements(By.XPATH, '//a[contains(@href, "mailto")]')
                email = 'Не указан' if len(email) == 0 else email[0].text
                print(self.inn_iter_count)
                data = [[inn_list[self.inn_iter_count], company_name[self.inn_iter_count], phone, email]]
                self.inn_iter_count += 1
                self.add_data_to_excel(data)
        except IndexError as e:
            driver.close()
            driver.quit()

            return None
        driver.close()
        driver.quit()

    @staticmethod
    def get_inn_list():
        workbook = load_workbook('./documents/inn.xlsx', data_only=True)
        sheet = workbook.active
        inn = list([int(row[0].value) for row in list(sheet)[1:]])
        company = list([row[1].value for row in list(sheet)[1:]])

        return inn, company

    @staticmethod
    def create_excel():
        df = pd.DataFrame(columns=[
            'ИНН',
            'НАИМЕНОВАНИЕ',
            'НОМЕР ТЕЛЕФОНА',
            'ЭЛЕКТРОННАЯ ПОЧТА',
        ])
        df.to_excel(f'./documents/inn_ready.xlsx', index=False)

    @staticmethod
    def add_data_to_excel(data):
        workbook = openpyxl.load_workbook('./documents/inn_ready.xlsx')
        sheet = workbook.active

        for row in data:
            sheet.append(row)

        workbook.save('./documents/inn_ready.xlsx')



parser = Parser()

